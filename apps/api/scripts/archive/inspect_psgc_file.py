"""
Inspect PSGC Excel file structure to understand its format.
"""
import sys
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed")
    sys.exit(1)

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../..'))
psgc_file = Path(project_root) / 'data' / 'PSGC-July-2025-Publication-Datafile.xlsx'

print(f"Reading: {psgc_file}")
wb = openpyxl.load_workbook(psgc_file, data_only=True)
ws = wb.active

print(f"\nSheet name: {ws.title}")
print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")

# Read first 20 rows to understand structure
print("\nFirst 20 rows:")
for idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
    if any(cell for cell in row if cell):
        print(f"\nRow {idx}:")
        for col_idx, cell in enumerate(row[:10], 1):  # First 10 columns
            if cell:
                print(f"  Col {col_idx}: {cell}")

















